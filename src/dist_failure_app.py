from __future__ import annotations

import json
import os
import re
import sqlite3
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
import pandas as pd
import requests
import streamlit as st

try:
    from src.manual_reloads import build_manual_reload_profile, manual_reload_profile, run_manual_reload_workflow
except ModuleNotFoundError:
    # Streamlit Cloud may execute this file as a script from src/, where "src" isn't importable as a package.
    from manual_reloads import build_manual_reload_profile, manual_reload_profile, run_manual_reload_workflow


SOURCE_WORKBOOK_PATH = Path(
    r"C:\Users\psalmon\OneDrive - Digital Realty\ITE Team - 2026 Network Refresh\0.LCM Patching\1.Remediation\1.2026\5.August\US-APAC\Philip Salmon\East devices for Dist failures-07-15-2026_v1.2.xlsx"
)
SEED_START_ROW = 91
SEED_END_ROW = 182
SEED_CSV_PATH = SOURCE_WORKBOOK_PATH.parent / 'east_dist_failures_rows_91_182_seed.csv'
REPORT_CSV_PATH = SOURCE_WORKBOOK_PATH.parent / 'east_dist_failures_rows_91_182_report.csv'
REPORT_JSON_PATH = SOURCE_WORKBOOK_PATH.parent / 'east_dist_failures_rows_91_182_report.json'
POSTCHECK_EXPORT_DIR = SOURCE_WORKBOOK_PATH.parent / 'east_dist_failures_postchecks'
DATABASE_PATH = Path('data/dist_failure_remediation.db')
GOLDEN_IMAGE_INDEX_URL = 'http://10.3.5.40:8500/'
AUTO_IMAGE_RULES = [
    {
        'patterns': ['WS-C2960X', 'C2960X'],
        'image_name': 'c2960x-universalk9-mz.152-7.E14.bin',
        'target_version': '15.2(7)E14',
        'platform_prefix': '2960',
        'label': '2960X',
    },
    {
        'patterns': ['IE-3300', 'IE3300', 'IE-3'],
        'image_name': 'ie3x00-universalk9.17.15.05.SPA.bin',
        'target_version': '17.15.05',
        'platform_prefix': 'IE',
        'label': 'IE3x00',
    },
    {
        'patterns': ['C9200L', 'C9200-'],
        'image_name': 'cat9k_lite_iosxe.17.15.05.SPA.bin',
        'target_version': '17.15.05',
        'platform_prefix': '9200',
        'label': 'Cat9k Lite',
    },
    {
        'patterns': ['C9300L'],
        'image_name': 'cat9k_lite_iosxe.17.15.05.SPA.bin',
        'target_version': '17.15.05',
        'platform_prefix': '9300',
        'label': 'Cat9k Lite',
    },
    {
        'patterns': ['C9500', 'C9410R'],
        'image_name': 'cat9k_iosxe.17.15.05.SPA.bin',
        'target_version': '17.15.05',
        'platform_prefix': '9500',
        'label': 'Cat9k Modular',
    },
    {
        'patterns': ['ISR4451', 'ISR44'],
        'image_name': 'isr4400-universalk9.17.12.07b.SPA.bin',
        'target_version': '17.12.07b',
        'platform_prefix': 'ISR44',
        'label': 'ISR4400',
    },
    {
        'patterns': ['ISR4331', 'ISR43'],
        'image_name': 'isr4300-universalk9.17.12.07b.SPA.bin',
        'target_version': '17.12.07b',
        'platform_prefix': 'ISR43',
        'label': 'ISR4300',
    },
    {
        'patterns': ['C8300'],
        'image_name': 'c8000be-universalk9.17.15.05.SPA.bin',
        'target_version': '17.15.05',
        'platform_prefix': '8300',
        'label': 'Catalyst 8300',
    },
]


class _LinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.hrefs: List[str] = []

    def handle_starttag(self, tag: str, attrs: List[tuple[str, str | None]]) -> None:
        if tag.lower() != 'a':
            return
        for name, value in attrs:
            if name.lower() == 'href' and value:
                self.hrefs.append(value)


def _normalize_header(value: Any) -> str:
    return re.sub(r'[^a-z0-9]+', '', str(value or '').strip().lower())


def _extract_report_date(file_name: str) -> str:
    match = re.search(r'(\d{2})-(\d{2})-(\d{4})', file_name)
    if not match:
        return datetime.now().strftime('%Y-%m-%d')
    month, day, year = match.groups()
    return f'{year}-{month}-{day}'


def _derive_site_code(device_name: str) -> str:
    parts = str(device_name or '').split('.')
    if len(parts) >= 2:
        return parts[1].upper()
    return ''


def _safe_file_part(value: str) -> str:
    text = re.sub(r'[^A-Za-z0-9._-]+', '_', str(value or '').strip())
    return text[:120] or 'unknown'


def _empty_seed_row() -> Dict[str, str]:
    return {
        'Device Name': '',
        'Device IP': '',
        'DC': '',
        'Date': '',
        'DNAC EST Time': '',
        'Platform': '',
        'Business / Non Business Hours': '',
        'Assigned to': '',
        'Clean Up Status': '',
        'Patching Status': '',
        'Patching Status 2': '',
        'Observed Platform': '',
        'Observed Version': '',
        'Observed Image': '',
        'System Image': '',
        'Image Present': '',
        'Needs Upgrade': '',
        'Boot Target': '',
        'Boot Command': '',
        'Workflow Status': 'Seeded',
        'Workflow Notes': '',
        'Target Image': '',
        'Target Version': '',
        'Profile Selection': '',
        'Source Row': '',
        'Source Workbook': '',
    }


def extract_seed_rows(workbook_path: Path, start_row: int = SEED_START_ROW, end_row: int = SEED_END_ROW) -> List[Dict[str, str]]:
    if not workbook_path.exists():
        raise FileNotFoundError(f'Workbook not found: {workbook_path}')

    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_map = {
        _normalize_header(worksheet.cell(row=1, column=column).value): column
        for column in range(1, worksheet.max_column + 1)
    }

    report_date = _extract_report_date(workbook_path.name)
    rows: List[Dict[str, str]] = []
    for row_number in range(start_row, min(end_row, worksheet.max_row) + 1):
        device_name = worksheet.cell(row=row_number, column=header_map['devicename']).value or ''
        device_ip = worksheet.cell(row=row_number, column=header_map['ipaddress']).value or ''
        if not str(device_name).strip() and not str(device_ip).strip():
            continue

        row = _empty_seed_row()
        row.update(
            {
                'Device Name': str(device_name).strip(),
                'Device IP': str(device_ip).strip(),
                'DC': _derive_site_code(str(device_name).strip()),
                'Date': report_date,
                'Platform': str(worksheet.cell(row=row_number, column=header_map['platform']).value or '').strip(),
                'Assigned to': str(worksheet.cell(row=row_number, column=header_map['assignedtech']).value or '').strip(),
                'Patching Status': str(worksheet.cell(row=row_number, column=header_map['osupdatestatus']).value or '').strip(),
                'Patching Status 2': str(worksheet.cell(row=row_number, column=header_map['newdistrubutionstatus']).value or '').strip(),
                'Workflow Notes': 'Imported from workbook; run Analyze to fetch current version and image over SSH.',
                'Source Row': str(row_number),
                'Source Workbook': str(workbook_path),
            }
        )
        rows.append(row)

    return rows


def _connect_db() -> sqlite3.Connection:
    DATABASE_PATH.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(DATABASE_PATH)
    connection.execute(
        '''
        CREATE TABLE IF NOT EXISTS remediation_devices (
            device_ip TEXT PRIMARY KEY,
            device_name TEXT NOT NULL,
            source_row INTEGER NOT NULL,
            payload_json TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        '''
    )
    return connection


def save_rows(rows: List[Dict[str, str]]) -> None:
    timestamp = datetime.now().isoformat(timespec='seconds')
    with _connect_db() as connection:
        connection.execute('DELETE FROM remediation_devices')
        connection.executemany(
            'INSERT INTO remediation_devices (device_ip, device_name, source_row, payload_json, updated_at) VALUES (?, ?, ?, ?, ?)',
            [
                (
                    str(row.get('Device IP') or ''),
                    str(row.get('Device Name') or ''),
                    int(str(row.get('Source Row') or '0') or '0'),
                    json.dumps(row),
                    timestamp,
                )
                for row in rows
            ],
        )


def load_rows() -> List[Dict[str, str]]:
    with _connect_db() as connection:
        cursor = connection.execute(
            'SELECT payload_json FROM remediation_devices ORDER BY source_row, device_name, device_ip'
        )
        return [json.loads(payload_json) for (payload_json,) in cursor.fetchall()]


def _report_rows(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    report_rows: List[Dict[str, str]] = []
    for row in rows:
        report_rows.append(
            {
                'Device Name': row.get('Device Name', ''),
                'Device IP': row.get('Device IP', ''),
                'Platform': row.get('Platform', ''),
                'Observed Platform': row.get('Observed Platform', ''),
                'Observed Version': row.get('Observed Version', ''),
                'Observed Image': row.get('Observed Image', row.get('System Image', '')),
                'System Image': row.get('System Image', ''),
                'Image Present': row.get('Image Present', ''),
                'Needs Upgrade': row.get('Needs Upgrade', ''),
                'Target Image': row.get('Target Image', ''),
                'Target Version': row.get('Target Version', ''),
                'Profile Selection': row.get('Profile Selection', ''),
                'Boot Target': row.get('Boot Target', ''),
                'Boot Validation Status': row.get('Boot Validation Status', ''),
                'Post Check Status': row.get('Post Check Status', ''),
                'Post Check Commands': row.get('Post Check Commands', ''),
                'Assigned to': row.get('Assigned to', ''),
                'Patching Status': row.get('Patching Status', ''),
                'Clean Up Status': row.get('Clean Up Status', ''),
                'Workflow Status': row.get('Workflow Status', ''),
                'Workflow Notes': row.get('Workflow Notes', ''),
                'Source Row': row.get('Source Row', ''),
            }
        )
    return report_rows


def write_outputs(rows: List[Dict[str, str]]) -> None:
    SEED_CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(rows).to_csv(SEED_CSV_PATH, index=False)

    report_rows = _report_rows(rows)
    pd.DataFrame(report_rows).to_csv(REPORT_CSV_PATH, index=False)
    REPORT_JSON_PATH.write_text(json.dumps(report_rows, indent=2), encoding='utf-8')

    POSTCHECK_EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    for existing_file in POSTCHECK_EXPORT_DIR.glob('*.txt'):
        existing_file.unlink()

    for row in rows:
        postcheck_output = str(row.get('Post Check Output') or '').strip()
        workflow_log = str(row.get('Workflow Log') or '').strip()
        if not postcheck_output and not workflow_log:
            continue

        device_name = _safe_file_part(str(row.get('Device Name') or row.get('Device IP') or 'unknown'))
        device_ip = _safe_file_part(str(row.get('Device IP') or 'unknown'))
        source_row = _safe_file_part(str(row.get('Source Row') or 'unknown'))
        file_path = POSTCHECK_EXPORT_DIR / f'{source_row}_{device_name}_{device_ip}.txt'

        sections = [
            f"Device Name: {row.get('Device Name', '')}",
            f"Device IP: {row.get('Device IP', '')}",
            f"Target Image: {row.get('Target Image', '')}",
            f"Target Version: {row.get('Target Version', '')}",
            f"Boot Target: {row.get('Boot Target', '')}",
            f"Boot Validation Status: {row.get('Boot Validation Status', '')}",
            f"Post Check Status: {row.get('Post Check Status', '')}",
            f"Workflow Status: {row.get('Workflow Status', '')}",
            '',
            '===== POST CHECK COMMANDS =====',
            str(row.get('Post Check Commands', '')).strip(),
            '',
            '===== POST CHECK OUTPUT =====',
            postcheck_output,
            '',
            '===== WORKFLOW LOG =====',
            workflow_log,
        ]
        file_path.write_text('\n'.join(sections).rstrip() + '\n', encoding='utf-8')


def ensure_seed_data() -> List[Dict[str, str]]:
    rows = load_rows()
    if rows:
        return rows
    if not SOURCE_WORKBOOK_PATH.exists():
        return []
    rows = extract_seed_rows(SOURCE_WORKBOOK_PATH)
    save_rows(rows)
    write_outputs(rows)
    return rows


def _refresh_seed_from_workbook() -> List[Dict[str, str]]:
    if not SOURCE_WORKBOOK_PATH.exists():
        raise FileNotFoundError(f'Workbook not found: {SOURCE_WORKBOOK_PATH}')
    rows = extract_seed_rows(SOURCE_WORKBOOK_PATH)
    save_rows(rows)
    write_outputs(rows)
    return rows


def _load_golden_images(index_url: str) -> List[str]:
    response = requests.get(index_url, timeout=10)
    response.raise_for_status()
    parser = _LinkParser()
    parser.feed(response.text)
    images: List[str] = []
    for href in parser.hrefs:
        name = href.strip().strip('/')
        if not name.lower().endswith('.bin'):
            continue
        images.append(name.replace('%20', ' '))
    return sorted(set(images))


def _catalog_lookup(golden_images: List[str]) -> Dict[str, str]:
    return {image.lower(): image for image in golden_images}


def _row_platform_text(row: Dict[str, str]) -> str:
    observed = str(row.get('Observed Platform') or '').strip()
    imported = str(row.get('Platform') or '').strip()
    return observed or imported


def _resolved_catalog_image(preferred_name: str, golden_images: List[str]) -> str | None:
    lookup = _catalog_lookup(golden_images)
    exact = lookup.get(preferred_name.lower())
    if exact:
        return exact
    prefix = preferred_name.split('.')[0].lower()
    for key, value in lookup.items():
        if key.startswith(prefix):
            return value
    return None


def _resolve_row_profile(
    row: Dict[str, str],
    golden_images: List[str],
    fallback_profile: Dict[str, str],
    use_automatic_image_selection: bool,
) -> tuple[Dict[str, str], str]:
    if not use_automatic_image_selection:
        return dict(fallback_profile), 'Manual fallback'

    platform_text = _row_platform_text(row).upper()
    for rule in AUTO_IMAGE_RULES:
        if not any(pattern in platform_text for pattern in rule['patterns']):
            continue
        resolved_image = _resolved_catalog_image(rule['image_name'], golden_images)
        if not resolved_image:
            return dict(fallback_profile), f"Fallback: catalog missing {rule['image_name']}"
        profile = build_manual_reload_profile(
            image_name=resolved_image,
            image_base_url=fallback_profile['image_base_url'],
            target_version=rule['target_version'],
            platform_prefix=rule['platform_prefix'],
        )
        return profile, f"Auto: {rule['label']}"

    return dict(fallback_profile), f"Fallback: no rule for {_row_platform_text(row) or 'Unknown platform'}"


def _apply_target_recommendations(
    rows: List[Dict[str, str]],
    golden_images: List[str],
    fallback_profile: Dict[str, str],
    use_automatic_image_selection: bool,
) -> List[Dict[str, str]]:
    annotated_rows: List[Dict[str, str]] = []
    for row in rows:
        updated = dict(row)
        profile, selection = _resolve_row_profile(updated, golden_images, fallback_profile, use_automatic_image_selection)
        updated['Target Image'] = profile['image_name']
        updated['Target Version'] = profile['target_version']
        updated['Profile Selection'] = selection
        annotated_rows.append(updated)
    return annotated_rows


def _run_analysis(
    execute_actions: bool,
    cleanup_old_images: bool,
    image_name: str,
    target_version: str,
    image_base_url: str,
    golden_images: List[str],
    use_automatic_image_selection: bool,
) -> List[Dict[str, str]]:
    rows = load_rows() or extract_seed_rows(SOURCE_WORKBOOK_PATH)
    fallback_profile = build_manual_reload_profile(
        image_name=image_name,
        image_base_url=image_base_url,
        target_version=target_version,
    )
    resolved_rows = _apply_target_recommendations(rows, golden_images, fallback_profile, use_automatic_image_selection)
    results: List[Dict[str, str]] = []
    for row in resolved_rows:
        profile = build_manual_reload_profile(
            image_name=row.get('Target Image') or fallback_profile['image_name'],
            image_base_url=image_base_url,
            target_version=row.get('Target Version') or fallback_profile['target_version'],
        )
        results.extend(
            run_manual_reload_workflow(
                [row],
                execute_actions=execute_actions,
                cleanup_old_images=cleanup_old_images,
                execute_reload=False,
                profile=profile,
            )
        )

    # Keep a dedicated observed image field alongside the existing system image field.
    for row in results:
        observed_image = str(row.get('Observed Image') or row.get('System Image') or '').strip()
        row['Observed Image'] = observed_image

    save_rows(results)
    write_outputs(results)
    return results


def _observed_summary_frame(rows: List[Dict[str, str]]) -> pd.DataFrame:
    summary_rows: List[Dict[str, str]] = []
    for row in rows:
        summary_rows.append(
            {
                'Device Name': row.get('Device Name', ''),
                'Device IP': row.get('Device IP', ''),
                'Observed Platform': row.get('Observed Platform', ''),
                'Observed Image': row.get('Observed Image', row.get('System Image', '')),
            }
        )
    return pd.DataFrame(summary_rows)


def _display_frame(rows: List[Dict[str, str]]) -> pd.DataFrame:
    return pd.DataFrame(_report_rows(rows))


def _has_real_env_value(name: str) -> bool:
    value = str(os.getenv(name) or '').strip()
    if not value:
        return False
    placeholders = {'your_api_key_here', 'your_emea_token_here', 'your_us_token_here', 'your_apac_token_here'}
    return value not in placeholders


def _readiness_checks(workbook_exists: bool, listing_error: str, rows: List[Dict[str, str]], image_base_url: str) -> List[Dict[str, str]]:
    rows_needing_action = sum(1 for row in rows if row.get('Needs Upgrade') == 'Yes')
    rows_analyzed = sum(1 for row in rows if row.get('Observed Version'))
    return [
        {
            'Check': 'Workbook available',
            'Status': 'Ready' if workbook_exists else 'Blocked',
            'Details': str(SOURCE_WORKBOOK_PATH),
        },
        {
            'Check': 'Golden image catalog reachable',
            'Status': 'Ready' if not listing_error else 'Blocked',
            'Details': image_base_url if not listing_error else listing_error,
        },
        {
            'Check': 'SSH username configured',
            'Status': 'Ready' if _has_real_env_value('DNA_USERNAME') else 'Blocked',
            'Details': 'DNA_USERNAME present in environment' if _has_real_env_value('DNA_USERNAME') else 'DNA_USERNAME is missing or placeholder',
        },
        {
            'Check': 'SSH password configured',
            'Status': 'Ready' if _has_real_env_value('DNA_PASSWORD') else 'Blocked',
            'Details': 'DNA_PASSWORD present in environment' if _has_real_env_value('DNA_PASSWORD') else 'DNA_PASSWORD is missing',
        },
        {
            'Check': 'Device analysis completed',
            'Status': 'Ready' if rows_analyzed > 0 else 'Warning',
            'Details': f'{rows_analyzed} rows have observed versions; {rows_needing_action} currently marked as needing action',
        },
    ]


def _readiness_ok(checks: List[Dict[str, str]]) -> bool:
    return not any(check['Status'] == 'Blocked' for check in checks)


def render_app() -> None:
    st.set_page_config(page_title='Distribution Failure Remediation', layout='wide')
    st.title('Distribution Failure Remediation')

    default_profile = manual_reload_profile()
    st.caption(f'Source workbook: {SOURCE_WORKBOOK_PATH}')
    st.caption(f'Initial database: {DATABASE_PATH.resolve()}')
    st.caption(f'Per-device post-check exports: {POSTCHECK_EXPORT_DIR}')
    st.caption(
        'Workflow actions reuse the existing SSH remediation logic for this non-reload phase: fetch current version/image, download image if needed, validate the boot statement, and optionally clean old images.'
    )

    rows = ensure_seed_data()

    try:
        golden_images = _load_golden_images(GOLDEN_IMAGE_INDEX_URL)
        listing_error = ''
    except Exception as exc:
        golden_images = [default_profile['image_name']]
        listing_error = str(exc)

    default_index = 0
    if default_profile['image_name'] in golden_images:
        default_index = golden_images.index(default_profile['image_name'])

    cfg1, cfg2, cfg3 = st.columns([2, 1, 2])
    selected_image = cfg1.selectbox('Fallback golden image', options=golden_images, index=default_index)
    target_version = cfg2.text_input('Fallback target version', value=default_profile['target_version'])
    image_base_url = cfg3.text_input('Golden image base URL', value=GOLDEN_IMAGE_INDEX_URL)
    use_automatic_image_selection = st.checkbox('Auto-select image per platform', value=True)
    selected_profile = build_manual_reload_profile(
        image_name=selected_image,
        image_base_url=image_base_url,
        target_version=target_version,
    )
    st.caption(
        f'Fallback image: {selected_profile["image_name"]} | Fallback version: {selected_profile["target_version"]} | Image URL: {selected_profile["image_url"]}'
    )
    st.caption('Automatic rules cover 2960X, IE3300, C9200/C9300L, C9500/C9410R, ISR4331/4451, and C8300. Unmatched rows use the fallback image above.')
    if listing_error:
        st.warning(f'Golden image listing could not be loaded dynamically: {listing_error}')

    rows = _apply_target_recommendations(rows, golden_images, selected_profile, use_automatic_image_selection)

    readiness_checks = _readiness_checks(SOURCE_WORKBOOK_PATH.exists(), listing_error, rows, image_base_url)
    readiness_ok = _readiness_ok(readiness_checks)
    with st.expander('Execution Readiness', expanded=True):
        st.dataframe(pd.DataFrame(readiness_checks), width='stretch', hide_index=True)
        if readiness_ok:
            st.caption('Environment checks passed for staging operations.')
        else:
            st.warning('Resolve blocked readiness checks before running staging actions.')

    c1, c2, c3 = st.columns(3)
    if c1.button('Refresh Seed from Workbook', use_container_width=True):
        try:
            rows = _refresh_seed_from_workbook()
            rows = _apply_target_recommendations(rows, golden_images, selected_profile, use_automatic_image_selection)
            save_rows(rows)
            write_outputs(rows)
            st.success('Seed database refreshed from workbook rows 91-182.')
        except FileNotFoundError as exc:
            st.error(str(exc))
    if c2.button('Analyze Version/Image Report', use_container_width=True):
        with st.spinner('Collecting show version and image state over SSH...'):
            rows = _run_analysis(
                execute_actions=False,
                cleanup_old_images=False,
                image_name=selected_image,
                target_version=target_version,
                image_base_url=image_base_url,
                golden_images=golden_images,
                use_automatic_image_selection=use_automatic_image_selection,
            )
        st.success('Version/image analysis complete.')

    execute_actions = st.checkbox('Execute image download, boot statement update, validation, and write memory', value=False)
    cleanup_old_images = st.checkbox('Clean up old images during execution', value=False)
    confirmation_checked = st.checkbox('I confirm this phase will modify boot settings and write memory, but will not reload devices', value=False)
    confirmation_text = st.text_input('Type STAGE to enable staging writes', value='')
    can_stage = readiness_ok and execute_actions and confirmation_checked and confirmation_text.strip().upper() == 'STAGE'
    if c3.button('Run Staging Workflow', use_container_width=True, type='primary', disabled=not can_stage):
        with st.spinner('Running remediation workflow against the current database...'):
            rows = _run_analysis(
                execute_actions=execute_actions,
                cleanup_old_images=cleanup_old_images,
                image_name=selected_image,
                target_version=target_version,
                image_base_url=image_base_url,
                golden_images=golden_images,
                use_automatic_image_selection=use_automatic_image_selection,
            )
        st.success('Staging workflow finished. No reloads were sent. Review the report below.')
    elif execute_actions and not can_stage:
        st.info('Staging writes remain disabled until readiness checks pass and the confirmation step is completed.')

    rows = load_rows()
    rows = _apply_target_recommendations(rows, golden_images, selected_profile, use_automatic_image_selection)
    for row in rows:
        row['Observed Image'] = str(row.get('Observed Image') or row.get('System Image') or '').strip()
    total = len(rows)
    observed = sum(1 for row in rows if row.get('Observed Version'))
    needs_upgrade = sum(1 for row in rows if row.get('Needs Upgrade') == 'Yes')
    cleanup_done = sum(1 for row in rows if row.get('Clean Up Status') == 'Completed')
    boot_validated = sum(1 for row in rows if row.get('Boot Validation Status') == 'Matched running-config')
    auto_matched = sum(1 for row in rows if str(row.get('Profile Selection') or '').startswith('Auto:'))
    postchecked = sum(1 for row in rows if row.get('Post Check Status') == 'Completed')
    m1, m2, m3, m4, m5, m6, m7 = st.columns(7)
    m1.metric('Seed Rows', total)
    m2.metric('Versions Collected', observed)
    m3.metric('Need Upgrade', needs_upgrade)
    m4.metric('Cleanup Completed', cleanup_done)
    m5.metric('Auto Matched', auto_matched)
    m6.metric('Boot Validated', boot_validated)
    m7.metric('Post Checks', postchecked)

    frame = _display_frame(rows)
    st.dataframe(frame, width='stretch', hide_index=True)

    st.subheader('Observed Platform and Image (All Devices)')
    st.dataframe(_observed_summary_frame(rows), width='stretch', hide_index=True)

    st.download_button(
        'Download Seed CSV',
        data=SEED_CSV_PATH.read_bytes() if SEED_CSV_PATH.exists() else b'',
        file_name=SEED_CSV_PATH.name,
        mime='text/csv',
        use_container_width=True,
    )
    st.download_button(
        'Download Report CSV',
        data=REPORT_CSV_PATH.read_bytes() if REPORT_CSV_PATH.exists() else b'',
        file_name=REPORT_CSV_PATH.name,
        mime='text/csv',
        use_container_width=True,
    )
    st.download_button(
        'Download Report JSON',
        data=REPORT_JSON_PATH.read_bytes() if REPORT_JSON_PATH.exists() else b'',
        file_name=REPORT_JSON_PATH.name,
        mime='application/json',
        use_container_width=True,
    )


if __name__ == '__main__':
    render_app()